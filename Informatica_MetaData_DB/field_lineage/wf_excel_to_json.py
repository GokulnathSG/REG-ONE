#!/usr/bin/env python3
"""Reusable Excel-to-JSON converter optimized for read-only processing.

Requires: openpyxl
Install:  pip install openpyxl

Examples:
  python excel_to_json.py input.xlsx output.json
  python excel_to_json.py input.xlsx output.json --pretty
  python excel_to_json.py input.xlsx output.jsonl --format jsonl
  python excel_to_json.py input.xlsx output.json --sheets "Expression" "Filter"
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def json_value(value: Any) -> Any:
    """Convert Excel/Python values to JSON-safe values without altering text."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def unique_headers(values: Iterable[Any]) -> list[str]:
    """Create non-empty, unique JSON property names from the first row."""
    result: list[str] = []
    counts: dict[str, int] = {}
    for position, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else ""
        if not base:
            base = f"column_{position}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def worksheet_records(ws, skip_blank_rows: bool = True) -> tuple[list[str], list[dict[str, Any]]]:
    rows = ws.iter_rows(values_only=True)
    try:
        headers = unique_headers(next(rows))
    except StopIteration:
        return [], []

    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        padded = tuple(row) + (None,) * max(0, len(headers) - len(row))
        values = padded[: len(headers)]
        if skip_blank_rows and all(value is None for value in values):
            continue
        record = {header: json_value(value) for header, value in zip(headers, values)}
        records.append(record)
    return headers, records


def convert_to_json(input_path: Path, output_path: Path, sheets: list[str] | None,
                    pretty: bool, include_nulls: bool, data_only: bool) -> None:
    wb = load_workbook(input_path, read_only=True, data_only=data_only)
    selected = sheets or wb.sheetnames
    missing = [name for name in selected if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"Sheet(s) not found: {', '.join(missing)}")

    payload: dict[str, Any] = {
        "source_file": input_path.name,
        "format_version": "1.0",
        "sheets": {},
    }

    for name in selected:
        headers, records = worksheet_records(wb[name])
        if not include_nulls:
            records = [{k: v for k, v in record.items() if v is not None} for record in records]
        payload["sheets"][name] = {
            "row_count": len(records),
            "columns": headers,
            "records": records,
        }

    with output_path.open("w", encoding="utf-8", newline="\n") as file:
        json.dump(
            payload,
            file,
            ensure_ascii=False,
            indent=2 if pretty else None,
            separators=None if pretty else (",", ":"),
            allow_nan=False,
        )
        file.write("\n")


def convert_to_jsonl(input_path: Path, output_path: Path, sheets: list[str] | None,
                     include_nulls: bool, data_only: bool) -> None:
    """Write one record per line for low-memory streaming and faster partial reads."""
    wb = load_workbook(input_path, read_only=True, data_only=data_only)
    selected = sheets or wb.sheetnames
    missing = [name for name in selected if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"Sheet(s) not found: {', '.join(missing)}")

    with output_path.open("w", encoding="utf-8", newline="\n") as file:
        for name in selected:
            headers, records = worksheet_records(wb[name])
            for record in records:
                if not include_nulls:
                    record = {k: v for k, v in record.items() if v is not None}
                item = {"sheet": name, "record": record}
                file.write(json.dumps(item, ensure_ascii=False, separators=(",", ":"), allow_nan=False))
                file.write("\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert every Excel worksheet to JSON or JSON Lines.")
    parser.add_argument("input", type=Path, help="Input .xlsx file")
    parser.add_argument("output", type=Path, help="Output .json or .jsonl file")
    parser.add_argument("--format", choices=("json", "jsonl"), default="json")
    parser.add_argument("--sheets", nargs="+", help="Optional sheet names; default is all sheets")
    parser.add_argument("--pretty", action="store_true", help="Indent JSON for readability (larger/slower)")
    parser.add_argument("--omit-nulls", action="store_true", help="Remove keys whose values are blank")
    parser.add_argument("--data-only", action="store_true", help="Use cached formula results instead of formulas")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.is_file():
        raise FileNotFoundError(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)

    if args.format == "jsonl":
        convert_to_jsonl(args.input, args.output, args.sheets, not args.omit_nulls, args.data_only)
    else:
        convert_to_json(args.input, args.output, args.sheets, args.pretty,
                        not args.omit_nulls, args.data_only)

    print(f"Created: {args.output}")


if __name__ == "__main__":
    main()
