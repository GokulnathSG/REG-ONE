#!/usr/bin/env python3
"""Create an Informatica field-lineage workbook from:
1) a direct Informatica export converted to hierarchical JSON, and
2) an Excel workbook (or its JSON conversion) containing transformation details.

Usage:
  python informatica_field_lineage.py direct.json transformations.xlsx -o Field_Lineage.xlsx
  python informatica_field_lineage.py direct.json input_2.json -o Field_Lineage.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict, deque
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Mapping", "Mapplet", "Session", "Source Table", "Source Field",
    "Target Table", "Target Field", "Transformation", "Transformation Type",
    "Transformation_Full_Lineage_path", "Implementation/Business logic", "Links",
]


def children(node: dict, tag: str | None = None) -> list[dict]:
    result = [x for x in node.get("children", []) if isinstance(x, dict)]
    return result if tag is None else [x for x in result if x.get("tag") == tag]


def walk(node: Any, tag: str | None = None) -> Iterable[dict]:
    if isinstance(node, dict):
        if tag is None or node.get("tag") == tag:
            yield node
        for value in node.values():
            yield from walk(value, tag)
    elif isinstance(node, list):
        for value in node:
            yield from walk(value, tag)


def attr(node: dict, key: str, default: str = "") -> str:
    value = node.get("attributes", {}).get(key, default)
    return "" if value is None else str(value)


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def load_direct_json(path: Path) -> dict:
    with path.open(encoding="utf-8-sig") as f:
        return json.load(f)


def excel_to_json(path: Path) -> dict:
    """Convert Excel values to the same {sheets:{name:{columns,records}}} shape."""
    wb = load_workbook(path, read_only=True, data_only=True)
    output = {"source_file": path.name, "format_version": "1.0", "sheets": {}}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            continue
        headers = [str(v).strip() if v is not None else f"COLUMN_{i+1}" for i, v in enumerate(raw_headers)]
        records = []
        for values in rows:
            if not any(v is not None and str(v).strip() for v in values):
                continue
            values = list(values) + [None] * (len(headers) - len(values))
            records.append(dict(zip(headers, values[:len(headers)])))
        output["sheets"][ws.title] = {"row_count": len(records), "columns": headers, "records": records}
    return output


def load_transform_data(path: Path, save_converted_json: Path | None = None) -> dict:
    if path.suffix.lower() == ".json":
        with path.open(encoding="utf-8-sig") as f:
            return json.load(f)
    data = excel_to_json(path)
    if save_converted_json:
        save_converted_json.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return data


def ordered_sessions(direct: dict) -> list[tuple[str, str]]:
    """Return (session instance name, session definition name) in workflow dependency order."""
    workflow = next(walk(direct, "WORKFLOW"), None)
    if not workflow:
        return []
    tasks = [n for n in children(workflow, "TASKINSTANCE") if norm(attr(n, "TASKTYPE")) == "session"]
    task_order = [attr(n, "NAME") for n in tasks]
    definition = {attr(n, "NAME"): attr(n, "TASKNAME") or attr(n, "NAME") for n in tasks}
    pos = {name: i for i, name in enumerate(task_order)}
    adj, indeg = defaultdict(list), {name: 0 for name in task_order}
    for link in children(workflow, "WORKFLOWLINK"):
        a, b = attr(link, "FROMTASK"), attr(link, "TOTASK")
        if a in indeg and b in indeg and b not in adj[a]:
            adj[a].append(b); indeg[b] += 1
    ready = [n for n in task_order if indeg[n] == 0]
    ready.sort(key=pos.get)
    result = []
    while ready:
        n = ready.pop(0); result.append(n)
        for m in sorted(adj[n], key=pos.get):
            indeg[m] -= 1
            if indeg[m] == 0:
                ready.append(m); ready.sort(key=pos.get)
    result += [n for n in task_order if n not in result]  # retain disconnected/cyclic tasks in document order
    return [(name, definition[name]) for name in result]


def transformation_index(data: dict) -> dict:
    """Index all transformation-detail rows, matching names case-insensitively."""
    idx = defaultdict(list)
    for sheet_name, sheet in data.get("sheets", {}).items():
        for seq, row in enumerate(sheet.get("records", [])):
            if not isinstance(row, dict):
                continue
            tname = row.get("TRANSFORMATION_NAME") or row.get("INSTANCE_NAME")
            mapping = row.get("MAPPING_NAME")
            session = row.get("SESSION_NAME") or row.get("SESSION_INSTANCE_NAME")
            if tname:
                copied = dict(row)
                copied["__sheet"] = sheet_name
                copied["__seq"] = seq
                idx[(norm(session), norm(mapping), norm(tname))].append(copied)
                idx[("", norm(mapping), norm(tname))].append(copied)
                idx[("", "", norm(tname))].append(copied)
    return idx


def logic_for(idx: dict, session: str, mapping: str, tname: str, ttype: str, port: str = "") -> str:
    rows = idx.get((norm(session), norm(mapping), norm(tname))) or idx.get(("", norm(mapping), norm(tname))) or idx.get(("", "", norm(tname))) or []
    if port:
        port_rows = [r for r in rows if norm(r.get("PORT_NAME")) == norm(port)]
        if port_rows:
            rows = port_rows
    typ = norm(ttype)
    if "expression" in typ:
        keys = ["EXPRESSION"]
    elif "lookup" in typ:
        keys = ["LOOKUP_CONDITION", "LOOKUP_TABLE_NAME", "LOOKUP_SQL_OVERRIDE"]
    elif "filter" in typ:
        keys = ["FILTER_CONDITION", "FILTER_EXPRESSION", "CONDITION"]
    else:
        # Preserve useful implementation metadata for other transformation categories.
        keys = [k for k in (rows[0].keys() if rows else []) if any(x in k.upper() for x in ("CONDITION", "EXPRESSION", "SQL_OVERRIDE", "TABLE_NAME"))]
    items, seen = [], set()
    for row in rows:
        port_name = row.get("PORT_NAME")
        for key in keys:
            value = row.get(key)
            if value is None or str(value).strip() == "":
                continue
            text = f"{port_name}: {key}={value}" if port_name and len(rows) > 1 else f"{key}={value}"
            if text not in seen:
                seen.add(text); items.append(text)
    return "\n".join(items)


def resolve_mapping(direct: dict, session_definition: str) -> tuple[str, dict | None]:
    sessions = {attr(n, "NAME"): n for n in walk(direct, "SESSION")}
    s = sessions.get(session_definition)
    mapping_name = attr(s, "MAPPINGNAME") if s else ""
    mappings = {attr(n, "NAME"): n for n in walk(direct, "MAPPING")}
    return mapping_name, mappings.get(mapping_name)


def mapping_rows(mapping: dict, mapping_name: str, session_name: str, logic_idx: dict) -> list[dict]:
    instances = children(mapping, "INSTANCE")
    inst_meta = {}
    for i, n in enumerate(instances):
        name = attr(n, "NAME")
        associated = children(n, "ASSOCIATED_SOURCE_INSTANCE")
        source_base = attr(associated[0], "NAME") if associated else ""
        inst_meta[name] = {
            "order": i, "name": name, "base": attr(n, "TRANSFORMATION_NAME") or name,
            "source_base": source_base,
            "type": attr(n, "TRANSFORMATION_TYPE") or attr(n, "TYPE"),
            "mapplet": attr(n, "TRANSFORMATION_NAME") if "mapplet" in norm(attr(n, "TYPE") + " " + attr(n, "TRANSFORMATION_TYPE")) else "",
        }
    connectors = children(mapping, "CONNECTOR")
    outgoing, incoming = defaultdict(list), defaultdict(list)
    for i, c in enumerate(connectors):
        a = (attr(c, "FROMINSTANCE"), attr(c, "FROMFIELD"))
        b = (attr(c, "TOINSTANCE"), attr(c, "TOFIELD"))
        edge = (b, i)
        outgoing[a].append(edge); incoming[b].append((a, i))
    for edges in outgoing.values():
        edges.sort(key=lambda x: (inst_meta.get(x[0][0], {}).get("order", 10**9), x[1]))

    # A valid lineage must terminate at a real Informatica target instance.
    # Starts may be source qualifiers, lookups, sequence generators, or constant-producing transformations.
    source_nodes = {node for node in outgoing if not incoming.get(node) or "source definition" in norm(inst_meta.get(node[0], {}).get("type"))}
    target_nodes = {node for node in incoming if "target definition" in norm(inst_meta.get(node[0], {}).get("type"))}
    paths = []
    max_depth = max(20, len(connectors) + 1)
    for start in sorted(source_nodes, key=lambda n: (inst_meta.get(n[0], {}).get("order", 10**9), n[0], n[1])):
        stack = [(start, [start], {start})]
        while stack:
            node, path, visited = stack.pop()
            if node in target_nodes:
                paths.append(path); continue
            nexts = outgoing.get(node, [])
            for nxt, _ in reversed(nexts):
                if nxt not in visited and len(path) < max_depth:
                    stack.append((nxt, path + [nxt], visited | {nxt}))

    rows = []
    for path in paths:
        src, tgt = path[0], path[-1]
        src_meta, tgt_meta = inst_meta.get(src[0], {}), inst_meta.get(tgt[0], {})
        source_table = src_meta.get("source_base") or src_meta.get("base", src[0])
        target_table = tgt_meta.get("base", tgt[0])
        full_path = " -> ".join(f"{inst_meta.get(i, {}).get('base', i)}.{f}" for i, f in path)
        mapplets = [inst_meta.get(i, {}).get("mapplet", "") for i, _ in path if inst_meta.get(i, {}).get("mapplet")]
        transformed = [(i, f) for i, f in path[1:-1] if not any(x in norm(inst_meta.get(i, {}).get("type")) for x in ("source definition", "target definition"))]
        if not transformed:
            transformed = [(src[0], src[1])]
        for tinst, tfield in transformed:
            meta = inst_meta.get(tinst, {})
            tname, ttype = meta.get("base", tinst), meta.get("type", "")
            rows.append({
                "Mapping": mapping_name, "Mapplet": ", ".join(dict.fromkeys(mapplets)), "Session": session_name,
                "Source Table": source_table, "Source Field": src[1], "Target Table": target_table, "Target Field": tgt[1],
                "Transformation": tname, "Transformation Type": ttype,
                "Transformation_Full_Lineage_path": full_path,
                "Implementation/Business logic": logic_for(logic_idx, session_name, mapping_name, tname, ttype, tfield),
                "Links": "",
                "__trans_order": meta.get("order", 10**9),
            })
    rows.sort(key=lambda r: (r["__trans_order"], r["Source Table"], r["Source Field"], r["Target Table"], r["Target Field"]))
    return rows


def add_cross_session_links(all_session_rows: list[list[dict]]) -> None:
    previous_targets = defaultdict(list)
    for session_rows in all_session_rows:
        for row in session_rows:
            key = (norm(row["Source Table"]), norm(row["Source Field"]))
            matches = previous_targets.get(key, [])
            if matches:
                row["Links"] = "\n".join(dict.fromkeys(
                    f'{p["Session"]}.{p["Target Table"]}.{p["Target Field"]}->{row["Session"]}.{row["Source Table"]}.{row["Source Field"]}'
                    for p in matches
                ))
        for row in session_rows:
            previous_targets[(norm(row["Target Table"]), norm(row["Target Field"]))].append(row)


def write_workbook(rows: list[dict], output: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Field_Lineage"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; ws.sheet_view.showGridLines = False
    widths = [28, 24, 30, 28, 24, 28, 24, 28, 22, 80, 80, 65]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(color="008000")  # imported/linked data
    ws.row_dimensions[1].height = 33
    wb.save(output)


def main() -> None:
    p = argparse.ArgumentParser(description="Generate Informatica field lineage in execution order.")
    p.add_argument("direct_json", type=Path, help="Hierarchical JSON converted from Informatica XML")
    p.add_argument("transformation_input", type=Path, help="Transformation-detail .xlsx or Excel-converted .json")
    p.add_argument("-o", "--output", type=Path, default=Path("Field_Lineage.xlsx"))
    p.add_argument("--save-converted-json", type=Path, help="When input 2 is Excel, optionally save its values as JSON")
    args = p.parse_args()

    direct = load_direct_json(args.direct_json)
    transform_data = load_transform_data(args.transformation_input, args.save_converted_json)
    logic_idx = transformation_index(transform_data)
    session_groups = []
    session_order = ordered_sessions(direct)
    for session_instance, session_definition in session_order:
        mapping_name, mapping = resolve_mapping(direct, session_definition)
        if not mapping:
            print(f"WARNING: mapping not found for session {session_instance} (definition={session_definition})")
            continue
        session_groups.append(mapping_rows(mapping, mapping_name, session_instance, logic_idx))
    add_cross_session_links(session_groups)
    rows = [row for group in session_groups for row in group]
    write_workbook(rows, args.output)
    print(f"Created {args.output} with {len(rows):,} lineage rows across {len(session_groups)} sessions.")
    print("Session execution order:")
    for i, (instance, definition) in enumerate(session_order, 1):
        print(f"  {i}. {instance} (definition: {definition})")


if __name__ == "__main__":
    main()
